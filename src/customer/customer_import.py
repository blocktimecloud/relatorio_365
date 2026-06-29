"""
Importação de clientes em massa via planilha (CSV ou Excel .xlsx).

A planilha deve conter um cabeçalho com as colunas esperadas. Colunas
obrigatórias e opcionais estão descritas em COLUNAS_OBRIGATORIAS / COLUNAS_OPCIONAIS.

O importador é tolerante a falhas: cada linha é processada de forma isolada;
linhas com erro (CNPJ inválido, cliente duplicado, campo obrigatório ausente)
são puladas e registradas em um relatório final, sem interromper a importação.
"""
from __future__ import annotations

import csv
from dataclasses import dataclass, field
from pathlib import Path

from sqlalchemy.orm import Session

from customer.customer_service import CustomerService
from customer.validators import validate_cnpj
from core.exceptions.customer import (
    CustomerAlreadyExistsException,
    InvalidCNPJException,
)
from core.logging.logger import logger


# ── Definição das colunas ────────────────────────────────────────────────
# Os nomes são normalizados (minúsculo, sem espaços nas bordas) na leitura,
# então o cabeçalho da planilha pode usar maiúsculas/minúsculas livremente.

COLUNAS_OBRIGATORIAS = [
    "razao_social",
    "name",          # nome fantasia
    "cnpj",
    "tenant_id",
    "client_id",
    "client_secret",
    "sharepoint_name",  # ex: 'duco' para duco.sharepoint.com
]

COLUNAS_OPCIONAIS = [
    "contact_email",
    "recipient_name",
    "recipient_email",
]

TODAS_AS_COLUNAS = COLUNAS_OBRIGATORIAS + COLUNAS_OPCIONAIS


# ── Resultado da importação ──────────────────────────────────────────────

@dataclass
class LinhaErro:
    linha: int          # número da linha na planilha (1-based, contando o cabeçalho)
    identificacao: str  # algo que ajude o usuário a achar a linha (nome ou cnpj)
    motivo: str


@dataclass
class ResultadoImportacao:
    criados: int = 0
    erros: list[LinhaErro] = field(default_factory=list)

    @property
    def total_processado(self) -> int:
        return self.criados + len(self.erros)

    def resumo(self) -> str:
        linhas = [
            "",
            "── Resultado da importação ─────────────────────────",
            f"  Clientes criados : {self.criados}",
            f"  Linhas com erro  : {len(self.erros)}",
            f"  Total processado : {self.total_processado}",
        ]
        if self.erros:
            linhas.append("")
            linhas.append("  Erros (linhas puladas):")
            for e in self.erros:
                linhas.append(f"    • Linha {e.linha} [{e.identificacao}]: {e.motivo}")
        linhas.append("")
        return "\n".join(linhas)


# ── Leitura das planilhas ────────────────────────────────────────────────

def _normalizar_cabecalho(campos: list[str]) -> list[str]:
    return [(c or "").strip().lower() for c in campos]


def _ler_csv(caminho: Path) -> list[dict]:
    # utf-8-sig lida com o BOM que o Excel costuma adicionar ao salvar CSV
    with open(caminho, newline="", encoding="utf-8-sig") as f:
        # detecta o separador (vírgula ou ponto-e-vírgula, comum no Brasil)
        amostra = f.read(2048)
        f.seek(0)
        try:
            dialect = csv.Sniffer().sniff(amostra, delimiters=",;")
        except csv.Error:
            dialect = csv.excel  # fallback: vírgula padrão
        reader = csv.reader(f, dialect)
        linhas = list(reader)

    if not linhas:
        return []

    cabecalho = _normalizar_cabecalho(linhas[0])
    registros = []
    for valores in linhas[1:]:
        # ignora linhas completamente vazias
        if not any((v or "").strip() for v in valores):
            continue
        registro = {col: (valores[i].strip() if i < len(valores) and valores[i] else "")
                    for i, col in enumerate(cabecalho)}
        registros.append(registro)
    return registros


def _ler_xlsx(caminho: Path) -> list[dict]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ImportError(
            "Leitura de .xlsx requer o pacote 'openpyxl'. "
            "Instale com: pip install openpyxl"
        ) from exc

    wb = load_workbook(caminho, read_only=True, data_only=True)
    ws = wb.active
    linhas = list(ws.iter_rows(values_only=True))
    wb.close()

    if not linhas:
        return []

    cabecalho = _normalizar_cabecalho([str(c) if c is not None else "" for c in linhas[0]])
    registros = []
    for valores in linhas[1:]:
        if not any(v is not None and str(v).strip() for v in valores):
            continue
        registro = {}
        for i, col in enumerate(cabecalho):
            v = valores[i] if i < len(valores) else None
            registro[col] = str(v).strip() if v is not None else ""
        registros.append(registro)
    return registros


def ler_planilha(caminho: str | Path) -> list[dict]:
    """
    Lê a planilha (CSV ou XLSX, detectado pela extensão) e retorna uma lista
    de dicionários {coluna: valor}. Levanta FileNotFoundError ou ValueError
    se o arquivo não existir ou o formato não for suportado.
    """
    caminho = Path(caminho)
    if not caminho.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {caminho}")

    ext = caminho.suffix.lower()
    if ext == ".csv":
        return _ler_csv(caminho)
    if ext in (".xlsx", ".xlsm"):
        return _ler_xlsx(caminho)
    raise ValueError(
        f"Formato não suportado: '{ext}'. Use .csv ou .xlsx."
    )


# ── Validação de cabeçalho ───────────────────────────────────────────────

def validar_colunas(registros: list[dict]) -> list[str]:
    """
    Retorna a lista de colunas obrigatórias que estão FALTANDO no cabeçalho.
    Lista vazia significa que o cabeçalho está completo.
    """
    if not registros:
        return list(COLUNAS_OBRIGATORIAS)
    presentes = set(registros[0].keys())
    return [c for c in COLUNAS_OBRIGATORIAS if c not in presentes]


# ── Importação ───────────────────────────────────────────────────────────

def importar_clientes(db: Session, registros: list[dict]) -> ResultadoImportacao:
    """
    Processa os registros lidos da planilha, criando um cliente por linha.
    Linhas com erro são puladas e registradas no resultado.
    """
    resultado = ResultadoImportacao()
    svc = CustomerService(db)

    for indice, registro in enumerate(registros, start=2):  # +2: linha 1 é cabeçalho
        # identificação amigável da linha para o relatório de erros
        ident = registro.get("name") or registro.get("razao_social") \
            or registro.get("cnpj") or "—"

        # 1. campos obrigatórios presentes e não-vazios
        faltando = [c for c in COLUNAS_OBRIGATORIAS if not registro.get(c)]
        if faltando:
            resultado.erros.append(LinhaErro(
                linha=indice,
                identificacao=ident,
                motivo=f"campos obrigatórios ausentes: {', '.join(faltando)}",
            ))
            continue

        # 2. CNPJ válido
        try:
            cnpj_limpo = validate_cnpj(registro["cnpj"])
        except InvalidCNPJException as e:
            resultado.erros.append(LinhaErro(
                linha=indice, identificacao=ident, motivo=str(e),
            ))
            continue

        # 3. cria o cliente (secret vai em texto puro; o ORM criptografa)
        try:
            svc.create(
                name=registro["name"],
                razao_social=registro["razao_social"],
                cnpj=cnpj_limpo,
                tenant_id=registro["tenant_id"],
                client_id=registro["client_id"],
                client_secret=registro["client_secret"],
                contact_email=registro.get("contact_email") or None,
                recipient_name=registro.get("recipient_name") or None,
                recipient_email=registro.get("recipient_email") or None,
                sharepoint_name=registro["sharepoint_name"],
            )
            resultado.criados += 1
            logger.info(f"Importação: cliente criado '{registro['name']}' (linha {indice})")
        except CustomerAlreadyExistsException as e:
            resultado.erros.append(LinhaErro(
                linha=indice, identificacao=ident, motivo=str(e),
            ))
        except Exception as e:  # qualquer outro erro não derruba a importação inteira
            resultado.erros.append(LinhaErro(
                linha=indice, identificacao=ident, motivo=f"erro inesperado: {e}",
            ))
            logger.exception(f"Importação: falha na linha {indice}")

    return resultado